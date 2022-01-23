#!BOM_by_Openpyxl.py

def ad_for_cell(lst):
    x1 , y1 = lst[0], lst[1]
    colL = gcl(y1)
    strAd = colL + str(x1)
    return strAd

def lst2_for_rng(lst):
    c1, c2 = lst[0], lst[1]
    c1Ad, c2Ad= ad_for_cell(c1),ad_for_cell(c2);
    return (c1Ad, c2Ad)

def num_for_cell(str1):
    xy = cfs(str1)
    col = cifs(xy[0])
    return (xy[1], col)

def fn_Set_rTL(sht, strA, bnT):    
    for i in range(1, sht.max_row + 1):
        for j in range(1, sht.max_column + 1):
            if bnT:
                if sht.cell(i, j).value == strA:
                    ansLst = (sht, [i, sht.max_column])
                    return ansLst
            else:
                if strA in sht.cell(i, j).value:
                    ansLst = (sht, [i, sht.max_column])
                    return ansLst
    return None            
                
def fn_match_R(rA, strA, bnT):
    sht, l1, l2 = rA[0], rA[1], rA[2]
    x1, y1, x2, y2 = l1[0], l1[1], l2[0], l2[1]
    for i in range(x1, x2+1):
        for j in range(y1, y2+1):
            if bnT:
                if sht.cell(i, j).value == strA:
                    return (sht, [i, j])
            else:
                if strA in sht.cell(i, j).value:
                    return (sht, [i, j])
            return None                  

def fn_Slice_R(rA, strA, bnT):
    sht, l1, l2 = rA[0],rA[1], rA[2]
    x1, y1, x2, y2 = l1[0], l1[1], l2[0], l2[1]
    for i in range(x1, x2+1):
        for j in range(y1, y2+1):
            if bnT:
                if sht.cell(i, j).value == strA:
                    break
            else:
                if strA in sht.cell(i, j).value:
                    break
    return (sht, [i + 1, j], [sht.max_row, j])    
                
def faUniq(lst, bnT):
    sizeSet = set()
    for i in range(0, len(lst)):
        strTrim = lst[i].strip()
        strTrim = strTrim.replace(" ", "")
        sizeSet.add(strTrim) 
    if bnT:            
        return sorted(sizeSet)
    else:
        return list(sizeSet)

def fn_R_to_Lst(rA, bnT):
    vl = []
    sht, l1, l2 = rA[0], rA[1], rA[2]
    x1, y1, x2, y2 = l1[0], l1[1], l2[0], l2[1]
    for i in range(x1, x2+2):
        for j in range(y1, y2+1):
            if bnT:
                if sht.cell(i, j).value != None:
                    vl.append(sht.cell(i, j).value)
            else:
                vl.append(sht.cell(i, j).value)            
    return vl

def fn_Fmla_arC(rTL, lstStr, bnT):
    sht, l1, l2 = rTL[0], rTL[1],rTL[2]
    x1, y1, x2, y2 = l1[0], l1[1], l2[0], l2[1] 
    vC = []
    for i in range(x1, x2+2):
        for j in range(y1, y2+1):      
            if fb_Str_OK(sht.cell(i, j).value, lstStr, bnT):
                vC.append(j)    
    return vC    

def fb_Str_OK(strA, lsts, bnT):
    if bnT:        
        for k in range(len(lsts)):
            if strA == lsts[k]:
                return True
    else:
        for k in range(len(lsts)):
            if lsts[k] in strA:
                return True
    return False

def fn_Make_nc(rA, lc):
    sht, l1, l2 = rA[0], rA[1], rA[2]
    x1, y1, x2, y2 = l1[0], l1[1], l2[0], l2[1] 
    strLs = []; vl = [];
    for i in range(x1, x2 + 1):    
        for k in range(len(lc)):      
            sht.cell(i, lc[k]).value = sht.cell(i, lc[k]).value.replace(' ', '')
            strLs.append(sht.cell(i, lc[k]).value)
        vl.append(chr(9671).join(strLs))
        strLs = []
    return sorted(set(vl))

def fn_wr_in_list(sht, irw, lst, filename):
    #sht._WorkbookChild__title ==> 워크시트 이름 리턴
    vL = []
    fnh = os.path.splitext(filename)[0]
    areaName = fnh.split('_')[0]
    julju = fnh.split('_')[1]
    for i in range(len(lst)):
        vL = lst[i].split(',')
        for k in range(1, 3):
            if k == 0:
                sht.cell(irw, k).value = irw - 2
            elif k ==1:
                sht.cell(irw, k).value = areaName
            elif k ==2:     
                sht.cell(irw, k).value = julju
        for j in range(len(vL)):
            sht.cell(row = irw, column = j + 4).value = vL[j] 
        irw += 1
    return irw 

def trim_Apply(rA):
    sht, x1, y1, x2, y2 = rA[0], rA[1][0], rA[1][1], rA[2][0], rA[2][1]
    k = 0  
    for i in range(x1, x2+1):
        for j in range(y1, y2+1):
            sht.cell(i, j).value = sht.cell(i, j).value.strip()
            

def write_Title(rA, ls):  
    sht, x1, y1, x2, y2 = rA[0], rA[1][0], rA[1][1], rA[2][0], rA[2][1]
    k = 0  
    for i in range(x1, x2+1):
        for j in range(y1, y2+1):
            sht.cell(i, j).value = ls[k]
            k +=1

def column_width_apply(sht):
    for column_cells in sht.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        sht.column_dimensions[column_cells[0].column].width = length

def as_text(value):
    if value is None:
        return ""
    return str(value)

'''
import os, openpyxl
strWF = 'C:\\Py\\UT_원본\\작업시트'
os.makedirs(strWF, exist_ok = True)   
strFolder = 'C:\\Py\\UT_원본'; strlst = [];
wt = openpyxl.Workbook(); irw = 3; sht2 = wt.active; sht2.title = 'BOM';
rWT = [sht2, 2]
lstT = ['도면번호', '공종', 'P-MARK', '재질', '규격', 'B', '길이', 'TOTAL', '단중(Kg/Set)', 'TOTAL(Kg)' , \
        'EACH(m2)', 'TOTAL(m2)']
lstW = ['번호', '부위', '절구부', 'Assembly No.', '부재명', 'P-MARK', '재질', '규격', '폭', '길이', \
       '수량', '단중', '중량' , '단위면적', '도장면적']          
for fileName in os.listdir(strFolder):
    if fileName.endswith('xlsx'): 
        strlst = [];
        wb = openpyxl.load_workbook(os.path.join(strFolder, fileName))
        sht = wb['INPUT_DATA'] 
        rTL = fn_Set_rTL(sht, "규격", True); rA = fn_Slice_R(rTL, "규격", True);
        trim_Apply(rTL)                 
        lstC = fn_Fmla_arC(rTL, lstT, True)        
        #2-Assembly, 3-부재명, 5-P-Mark, 6-재질, 7-규격, 10-폭, 15, 길이, 17-수량, 19-단중, 20-중량, 21 -단위면적, 22-면적
        strlst = fn_Make_nc(rA, lstC)
        irw = fn_wr_in_list(sht2, irw, strlst, fileName)
        wb.close()
write_Title(rWT, lstW, 1)        
wt.save(os.path.join(strWF, 'Consolidated_BOM.xlsx')) 
wt.close()   
'''
import os, openpyxl
from openpyxl.utils.cell import coordinate_from_string as cfs, column_index_from_string as cifs
from openpyxl.utils import get_column_letter as gcl
wt = openpyxl.Workbook(); irw = 3; 
sh = wt.active; c1 = 'A1'; c2 = 'C1'
l1, l2 = num_for_cell(c1), num_for_cell(c2)
write_Title([sh, l1, l2], [1,2,3])
wt.save('D:\\연습.xlsx')